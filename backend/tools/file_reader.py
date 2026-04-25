from __future__ import annotations

import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Any
from xml.etree.ElementTree import ParseError

import pandas as pd
from openpyxl import load_workbook

_NS = "urn:schemas-microsoft-com:office:spreadsheet"

SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".csv", ".xls"}


def detect_format(filepath: Path) -> str:
    ext = filepath.suffix.lower()
    if ext == ".csv":
        return "csv"
    if ext == ".xlsm":
        return "xlsm"
    if ext == ".xlsx":
        return "xlsx"
    if ext == ".xls":
        with open(filepath, "rb") as fh:
            header = fh.read(2)
        return "xml_spreadsheet" if header == b"<?" else "xls_binary"
    raise ValueError(f"Unsupported extension: {ext}")


def read_file(filepath: Path) -> pd.DataFrame:
    """Return a RAW DataFrame with no header promotion.

    Row indices are 0-indexed from the first cell row of the sheet so they
    align with what `read_raw_cells()` sees. Header promotion happens later
    in `normalizer.apply_plan()` based on the DiscoveryPlan's
    `header_row_index`. Column names are integer positions (0, 1, 2, ...).
    """
    fmt = detect_format(filepath)
    if fmt in ("xlsx", "xlsm"):
        return pd.read_excel(filepath, engine="openpyxl", header=None)
    if fmt == "csv":
        return pd.read_csv(filepath, header=None)
    if fmt == "xls_binary":
        return pd.read_excel(filepath, engine="xlrd", header=None)
    if fmt == "xml_spreadsheet":
        # XML spreadsheet path keeps its own first-non-empty-row header
        # detection for now. NetSuite is the only producer of this format
        # and it rarely has a metadata preamble — revisit if needed.
        return _read_xml_spreadsheet(filepath)
    raise ValueError(f"Unknown format: {fmt}")


def read_raw_cells(
    filepath: Path,
    max_rows: int = 100,
    max_cols: int = 10,
) -> list[dict[str, Any]]:
    """Return raw rows with visual formatting flags for the Discovery sample.

    Each dict has:
      - row_index (int): 0-indexed position in the sheet
      - values (list): cell values across the first max_cols columns
      - is_bold (bool): any cell in the row is bold
      - indent_level (float): max indent across the row (0.0 if none)
      - is_merged (bool): any cell in the row is inside a merged range

    xlsx / xlsm: full formatting via openpyxl cell API.
    csv / xml / xls: formatting unavailable — flags default to
    False / 0.0 / False. Values are still captured so Discovery can
    still map columns; hierarchy recovery just has less signal.
    """
    fmt = detect_format(filepath)
    if fmt in ("xlsx", "xlsm"):
        return _read_raw_cells_openpyxl(filepath, max_rows, max_cols)
    return _read_raw_cells_fallback(filepath, fmt, max_rows, max_cols)


def _read_raw_cells_openpyxl(
    filepath: Path,
    max_rows: int,
    max_cols: int,
) -> list[dict[str, Any]]:
    # read_only=False is required to access font/alignment/merged ranges;
    # data_only=True evaluates formulas to their cached values.
    wb = load_workbook(filepath, data_only=True, read_only=False)
    ws = wb.active
    merged_ranges = list(ws.merged_cells.ranges)

    rows_out: list[dict[str, Any]] = []
    iterator = ws.iter_rows(min_row=1, max_row=max_rows, min_col=1, max_col=max_cols)
    for idx, row in enumerate(iterator):
        values: list[Any] = []
        is_bold = False
        indent_level = 0.0
        is_merged = False
        for cell in row:
            values.append(cell.value)
            font = cell.font
            if font is not None and font.bold:
                is_bold = True
            alignment = cell.alignment
            if alignment is not None and alignment.indent:
                indent_level = max(indent_level, float(alignment.indent))
            if not is_merged:
                coord = cell.coordinate
                for rng in merged_ranges:
                    if coord in rng:
                        is_merged = True
                        break
        rows_out.append(
            {
                "row_index": idx,
                "values": values,
                "is_bold": is_bold,
                "indent_level": indent_level,
                "is_merged": is_merged,
            }
        )
    wb.close()
    return rows_out


def _read_raw_cells_fallback(
    filepath: Path,
    fmt: str,
    max_rows: int,
    max_cols: int,
) -> list[dict[str, Any]]:
    if fmt == "csv":
        df = pd.read_csv(filepath, header=None, nrows=max_rows)
    elif fmt == "xls_binary":
        df = pd.read_excel(filepath, engine="xlrd", header=None, nrows=max_rows)
    elif fmt == "xml_spreadsheet":
        df = _read_xml_spreadsheet(filepath)
        df = df.head(max_rows)
    else:
        raise ValueError(f"Unsupported format for raw cell read: {fmt}")

    df = df.iloc[:, :max_cols]
    rows_out: list[dict[str, Any]] = []
    for idx, (_, row) in enumerate(df.iterrows()):
        values = [None if pd.isna(v) else v for v in row.tolist()]
        rows_out.append(
            {
                "row_index": idx,
                "values": values,
                "is_bold": False,
                "indent_level": 0.0,
                "is_merged": False,
            }
        )
    return rows_out


def _read_xml_spreadsheet(filepath: Path) -> pd.DataFrame:
    try:
        tree = ET.parse(filepath)
        root = tree.getroot()
        rows: list[list] = []

        for worksheet in root.findall(f"{{{_NS}}}Worksheet"):
            table = worksheet.find(f"{{{_NS}}}Table")
            if table is None:
                continue
            for row_elem in table.findall(f"{{{_NS}}}Row"):
                cells = row_elem.findall(f"{{{_NS}}}Cell")
                row_data = []
                for cell in cells:
                    data = cell.find(f"{{{_NS}}}Data")
                    if data is None:
                        row_data.append(None)
                        continue
                    type_attr = data.get(f"{{{_NS}}}Type", "String")
                    if type_attr == "Number":
                        row_data.append(float(data.text) if data.text else None)
                    else:
                        row_data.append(data.text)
                rows.append(row_data)
            break  # first worksheet only

        if not rows:
            return pd.DataFrame()

        # Skip leading rows that have no ss:Index attr on any cell (metadata heuristic).
        # The first row where all cells have plausible string content is the header.
        header_idx = 0
        for i, row in enumerate(rows):
            if any(v is not None for v in row):
                header_idx = i
                break

        header = rows[header_idx]
        data = rows[header_idx + 1 :]
        return pd.DataFrame(data, columns=header)

    except ParseError:
        # Some NetSuite exports are misnamed real .xls binaries — fall back to xlrd.
        return pd.read_excel(filepath, engine="xlrd")
