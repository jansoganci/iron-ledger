"""Excel export for IronLedger close packages.

Builds a 3-sheet .xlsx workbook from consolidated monthly data:
  Sheet 1 — Consolidated P&L   (one row per account, totals by category)
  Sheet 2 — Reconciliations     (cross-source discrepancies with severity)
  Sheet 3 — Source Breakdown    (per-account, per-file amounts)

Called by GET /report/{company_id}/{period}/export.xlsx.
Returns raw bytes; caller sets Content-Disposition header.
"""

from __future__ import annotations

import io
from datetime import date

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Palette
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")  # dark navy
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_SUBHEADER_FILL = PatternFill("solid", fgColor="D6E4F0")  # light blue
_SEVERITY_FILLS = {
    "high": PatternFill("solid", fgColor="FFCCCC"),
    "medium": PatternFill("solid", fgColor="FFE5B4"),
    "low": PatternFill("solid", fgColor="FFFFCC"),
}
_CURRENCY_FMT = "#,##0.00"
_PCT_FMT = "0.0%"


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def build_close_package(
    entries: list[dict],
    reconciliations: list[dict] | None,
    period: date,
    company_name: str,
) -> bytes:
    """Return raw .xlsx bytes for the close package workbook.

    Args:
        entries: list of dicts with keys: account, category, amount, source_breakdown
        reconciliations: list of ReconciliationItem dicts (may be None/empty)
        period: the reporting period
        company_name: shown in the header row
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default Sheet

    _build_pl_sheet(wb, entries, period, company_name)
    _build_reconciliation_sheet(wb, reconciliations or [], period)
    _build_source_breakdown_sheet(wb, entries, period)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Sheet 1 — Consolidated P&L
# ---------------------------------------------------------------------------

_CATEGORY_ORDER = ["REVENUE", "COGS", "OPEX", "G&A", "R&D", "OTHER_INCOME", "OTHER"]


def _build_pl_sheet(
    wb: openpyxl.Workbook,
    entries: list[dict],
    period: date,
    company_name: str,
) -> None:
    ws = wb.create_sheet("Consolidated P&L")

    # Title row
    ws.append([f"{company_name} — Consolidated P&L — {period.strftime('%B %Y')}"])
    _style_row(ws, 1, font=Font(bold=True, size=13))
    ws.merge_cells("A1:D1")

    # Header
    ws.append(["Account", "Category", "Amount ($)", "Sources"])
    _style_header_row(ws, 2)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 40

    # Group by category in prescribed order
    by_cat: dict[str, list[dict]] = {}
    for e in entries:
        cat = e.get("category", "OTHER")
        by_cat.setdefault(cat, []).append(e)

    row_num = 3
    for cat in _CATEGORY_ORDER:
        cat_entries = by_cat.get(cat, [])
        if not cat_entries:
            continue

        # Category sub-header
        ws.append([cat, "", "", ""])
        for col in range(1, 5):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = _SUBHEADER_FILL
            cell.font = Font(bold=True, size=10)
        row_num += 1

        cat_total = 0.0
        for e in sorted(cat_entries, key=lambda x: x["account"]):
            amount = float(e.get("amount", 0))
            cat_total += amount
            breakdown = e.get("source_breakdown") or []
            source_labels = ", ".join(
                f"{b['source_file']} ${b['amount']:,.0f}" for b in breakdown
            )
            ws.append([e["account"], cat, amount, source_labels])
            amt_cell = ws.cell(row=row_num, column=3)
            amt_cell.number_format = _CURRENCY_FMT
            row_num += 1

        # Category total
        ws.append(["", f"Total {cat}", cat_total, ""])
        total_cell = ws.cell(row=row_num, column=3)
        total_cell.number_format = _CURRENCY_FMT
        ws.cell(row=row_num, column=2).font = Font(bold=True)
        total_cell.font = Font(bold=True)
        row_num += 1

    ws.freeze_panes = "A3"


# ---------------------------------------------------------------------------
# Sheet 2 — Reconciliations
# ---------------------------------------------------------------------------


def _build_reconciliation_sheet(
    wb: openpyxl.Workbook,
    reconciliations: list[dict],
    period: date,
) -> None:
    ws = wb.create_sheet("Reconciliations")

    ws.append([f"Cross-Source Reconciliation — {period.strftime('%B %Y')}"])
    _style_row(ws, 1, font=Font(bold=True, size=13))
    ws.merge_cells("A1:G1")

    if not reconciliations:
        ws.append(["No cross-source discrepancies detected for this period."])
        return

    headers = [
        "Account",
        "Category",
        "GL Amount ($)",
        "Dept/Source Total ($)",
        "Delta ($)",
        "Severity",
        "Classification",
    ]
    ws.append(headers)
    _style_header_row(ws, 2)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 28

    row_num = 3
    for item in sorted(reconciliations, key=lambda x: -abs(x.get("delta", 0))):
        severity = item.get("severity", "low")
        row = [
            item.get("account", ""),
            item.get("category", ""),
            item.get("gl_amount"),
            item.get("non_gl_total"),
            item.get("delta"),
            severity.upper(),
            (item.get("classification") or "").replace("_", " ").title(),
        ]
        ws.append(row)
        fill = _SEVERITY_FILLS.get(severity)
        for col in range(1, 8):
            cell = ws.cell(row=row_num, column=col)
            if fill:
                cell.fill = fill
            if col in (3, 4, 5):
                cell.number_format = _CURRENCY_FMT
        row_num += 1

    ws.freeze_panes = "A3"

    # Sources detail block below
    row_num += 1
    ws.cell(row=row_num, column=1).value = "Source Detail"
    ws.cell(row=row_num, column=1).font = Font(bold=True, size=11)
    row_num += 1

    src_headers = ["Account", "Source File", "Amount ($)", "Row Count"]
    for col, h in enumerate(src_headers, 1):
        cell = ws.cell(row=row_num, column=col)
        cell.value = h
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    row_num += 1

    for item in reconciliations:
        for src in item.get("sources", []):
            ws.cell(row=row_num, column=1).value = item.get("account", "")
            ws.cell(row=row_num, column=2).value = src.get("source_file", "")
            amt_cell = ws.cell(row=row_num, column=3)
            amt_cell.value = src.get("amount")
            amt_cell.number_format = _CURRENCY_FMT
            ws.cell(row=row_num, column=4).value = src.get("row_count")
            row_num += 1


# ---------------------------------------------------------------------------
# Sheet 3 — Source Breakdown
# ---------------------------------------------------------------------------


def _build_source_breakdown_sheet(
    wb: openpyxl.Workbook,
    entries: list[dict],
    period: date,
) -> None:
    ws = wb.create_sheet("Source Breakdown")

    ws.append([f"Source Breakdown — {period.strftime('%B %Y')}"])
    _style_row(ws, 1, font=Font(bold=True, size=13))
    ws.merge_cells("A1:E1")

    headers = ["Account", "Category", "Source File", "Amount ($)", "Row Count"]
    ws.append(headers)
    _style_header_row(ws, 2)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 12

    row_num = 3
    for e in sorted(
        entries, key=lambda x: (x.get("category", ""), x.get("account", ""))
    ):
        breakdown = e.get("source_breakdown") or []
        if not breakdown:
            # Single-file run: write one row with filename as source
            ws.append(
                [
                    e["account"],
                    e.get("category", ""),
                    e.get("source_file", "—"),
                    float(e.get("amount", 0)),
                    "—",
                ]
            )
            ws.cell(row=row_num, column=4).number_format = _CURRENCY_FMT
            row_num += 1
        else:
            for src in breakdown:
                ws.append(
                    [
                        e["account"],
                        e.get("category", ""),
                        src.get("source_file", ""),
                        src.get("amount", 0),
                        src.get("row_count", ""),
                    ]
                )
                ws.cell(row=row_num, column=4).number_format = _CURRENCY_FMT
                row_num += 1

    ws.freeze_panes = "A3"


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------


def _style_header_row(ws, row_num: int) -> None:
    for col in range(1, ws.max_column + 2):
        cell = ws.cell(row=row_num, column=col)
        if cell.value is not None or col <= 7:
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center")


def _style_row(ws, row_num: int, font: Font | None = None) -> None:
    for cell in ws[row_num]:
        if font:
            cell.font = font
