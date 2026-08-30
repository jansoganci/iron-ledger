"""Excel export: coverage rows render as INFO / Not compared."""

from __future__ import annotations

from datetime import date
from io import BytesIO

import openpyxl

from backend.tools.excel_export import build_close_package


def test_coverage_row_exports_as_info_not_compared() -> None:
    raw = build_close_package(
        entries=[
            {
                "account": "Sales",
                "category": "REVENUE",
                "amount": 100.0,
                "source_breakdown": [],
            }
        ],
        reconciliations=[
            {
                "account": "Advertising — Meta",
                "category": "OPEX",
                "gl_amount": 8200.0,
                "non_gl_total": 0.0,
                "delta": -8200.0,
                "severity": "high",
                "classification": None,
                "card_kind": "coverage",
                "hints": {"is_gl_only": True},
                "sources": [],
            }
        ],
        period=date(2026, 3, 1),
        company_name="Vandelay",
    )
    wb = openpyxl.load_workbook(BytesIO(raw))
    ws = wb["Reconciliations"]
    accounts = {
        ws.cell(row=r, column=1).value: r
        for r in range(3, ws.max_row + 1)
    }
    row = accounts["Advertising — Meta"]
    assert ws.cell(row=row, column=6).value == "INFO"
    assert ws.cell(row=row, column=7).value == "Not compared"
