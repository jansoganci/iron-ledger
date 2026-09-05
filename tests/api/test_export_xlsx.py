"""The Excel export endpoint must actually return a workbook.

Regression for a 403 that every user hit on every request, found only by
driving the live HTTP path. The handler passed its own company check, then
called `get_by_owner(jwt_company_id)` — handing a COMPANY id to a lookup that
takes an OWNER id. Nothing matched, `get_by_owner` raises RLSForbiddenError on
an empty result, and that surfaced as 403. The endpoint had never worked.

The unit tests could not have caught it: they exercise `build_close_package`
directly and never reach the route. These tests go through the route.
"""

from __future__ import annotations

import io
from datetime import date
from unittest.mock import MagicMock, patch

import openpyxl
import pytest
from fastapi.testclient import TestClient

from backend.api.auth import (
    _company_cache,
    get_cached_company,
    get_company_id,
    get_current_user,
)
from backend.domain.entities import Report
from backend.main import app

USER_ID = "user-export-1"
COMPANY_ID = "co-export-1"
PERIOD = date(2026, 3, 1)

client = TestClient(app, raise_server_exceptions=False)


def _company_row() -> dict:
    return {
        "id": COMPANY_ID,
        "name": "Redhawk Alarm & Security LLC",
        "sector": "security",
        "currency": "USD",
        "monthly_revenue_band": "under_100k",
    }


@pytest.fixture(autouse=True)
def _override_auth():
    """Snapshot and restore, rather than set and pop.

    Two sibling modules (test_confirm_mappings, test_account_mapper_flow)
    install dependency_overrides at import time and never remove them, so the
    override map is whatever ran first. Overriding get_company_id explicitly
    makes these tests order-independent; restoring the exact snapshot means
    popping ours does not delete theirs.
    """
    saved = dict(app.dependency_overrides)
    app.dependency_overrides[get_current_user] = lambda: USER_ID
    app.dependency_overrides[get_company_id] = lambda: COMPANY_ID
    app.dependency_overrides[get_cached_company] = _company_row
    _company_cache[USER_ID] = (_company_row(), float("inf"))
    yield
    app.dependency_overrides.clear()
    app.dependency_overrides.update(saved)
    _company_cache.pop(USER_ID, None)


def _report() -> Report:
    return Report(
        id="report-1",
        company_id=COMPANY_ID,
        period=PERIOD,
        summary="Service Revenue shows a gap of 3 accounts totaling 285.00.",
        anomaly_count=1,
        error_count=0,
        reconciliations=[
            {
                "account": "Service Revenue",
                "category": "REVENUE",
                "gl_amount": 3540.0,
                "non_gl_total": 3825.0,
                "delta": 285.0,
                "delta_pct": 0.0805,
                "severity": "medium",
                "card_kind": "exception",
                "classification": "stale_reference",
                "sources": [
                    {
                        "source_file": "redhawk_contracts_mar_2026.xlsx",
                        "amount": 3825.0,
                        "row_count": 1,
                    }
                ],
            }
        ],
    )


class _Entry:
    account_id = "acct-1"
    actual_amount = 3540.0
    source_file = "redhawk_gl_mar_2026.xlsx"
    source_breakdown = [
        {
            "source_file": "redhawk_gl_mar_2026.xlsx",
            "amount": 3540.0,
            "row_count": 1,
        }
    ]


def _patched_repos():
    reports = MagicMock()
    reports.get.return_value = _report()
    entries = MagicMock()
    entries.list_for_period.return_value = [_Entry()]
    accounts = MagicMock()
    accounts.get_accounts_by_id.return_value = {
        "acct-1": {"name": "Service Revenue", "category": "REVENUE"}
    }
    return reports, entries, accounts


def test_export_returns_a_real_workbook_not_403() -> None:
    """The bug in one assertion: this returned 403 for every user."""
    reports, entries, accounts = _patched_repos()
    with patch("backend.api.routes.get_reports_repo", return_value=reports), patch(
        "backend.api.routes.get_entries_repo", return_value=entries
    ), patch("backend.api.routes.get_accounts_repo", return_value=accounts):
        resp = client.get(f"/report/{COMPANY_ID}/2026-03-01/export.xlsx")

    assert resp.status_code == 200, f"expected 200, got {resp.status_code}"
    assert "spreadsheetml" in resp.headers["content-type"]
    assert "attachment" in resp.headers["content-disposition"]

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert len(wb.sheetnames) == 3, f"expected 3 sheets, got {wb.sheetnames}"


def test_export_never_looks_a_company_up_by_its_own_id_as_an_owner() -> None:
    """Pins the specific mistake.

    `get_by_owner` takes an owner id. If the handler ever calls it with the
    company id again, the endpoint 403s for everyone — so assert it is not
    called that way at all.
    """
    reports, entries, accounts = _patched_repos()
    companies = MagicMock()
    with patch("backend.api.routes.get_reports_repo", return_value=reports), patch(
        "backend.api.routes.get_entries_repo", return_value=entries
    ), patch("backend.api.routes.get_accounts_repo", return_value=accounts), patch(
        "backend.api.routes.get_companies_repo", return_value=companies
    ):
        resp = client.get(f"/report/{COMPANY_ID}/2026-03-01/export.xlsx")

    assert resp.status_code == 200
    for call in companies.get_by_owner.call_args_list:
        assert COMPANY_ID not in call.args, "company id passed to get_by_owner"


def test_export_still_forbids_another_companys_report() -> None:
    """The fix must not loosen the tenancy check."""
    reports, entries, accounts = _patched_repos()
    with patch("backend.api.routes.get_reports_repo", return_value=reports), patch(
        "backend.api.routes.get_entries_repo", return_value=entries
    ), patch("backend.api.routes.get_accounts_repo", return_value=accounts):
        resp = client.get("/report/some-other-company/2026-03-01/export.xlsx")

    assert resp.status_code == 403


def test_export_404s_when_the_period_has_no_report() -> None:
    reports, entries, accounts = _patched_repos()
    reports.get.return_value = None
    with patch("backend.api.routes.get_reports_repo", return_value=reports), patch(
        "backend.api.routes.get_entries_repo", return_value=entries
    ), patch("backend.api.routes.get_accounts_repo", return_value=accounts):
        resp = client.get(f"/report/{COMPANY_ID}/2026-03-01/export.xlsx")

    assert resp.status_code == 404
