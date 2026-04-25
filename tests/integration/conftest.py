"""In-memory fixture bytes for Phase 4 integration tests.

No files checked into tests/fixtures/ — openpyxl + BytesIO synthesizes
everything at collection time. Keeps the repo slim and makes fixture
intent readable inline.
"""

from __future__ import annotations

import io
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

DRONE_CLEAN_PATH = Path("docs/demo_data/Drone Inc - Mar 26.xlsx")


@pytest.fixture
def drone_qbo_clean_bytes() -> bytes:
    """The real DRONE March P&L — our one piece of production-shape ground truth."""
    return DRONE_CLEAN_PATH.read_bytes()


@pytest.fixture
def flat_csv_bytes() -> bytes:
    """Trivial CSV — header at row 0, no metadata, no hierarchy."""
    return (
        "Account,Amount,Date\n"
        "Revenue,100000,2026-03-31\n"
        "COGS,-40000,2026-03-31\n"
        "OpEx,-35000,2026-03-31\n"
    ).encode("utf-8")


@pytest.fixture
def no_hierarchy_xlsx_bytes() -> bytes:
    """xlsx with headers + plain line items — no bold, no indent, no banners."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Account", "Amount"])
    ws.append(["Product Revenue", 45000])
    ws.append(["Service Revenue", 32000])
    ws.append(["Component Cost", -18000])
    ws.append(["Shipping", -3000])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def pii_laced_xlsx_bytes() -> bytes:
    """xlsx whose Memo column carries SSN, email, and credit-card patterns.

    Used to verify end-to-end that sanitize_sample redacts before any
    cell value reaches the LLM. The three PII shapes are all ones our
    regexes catch (hyphenated SSN, ASCII email, 4-4-4-4 CC).
    """
    wb = Workbook()
    ws = wb.active
    ws.append(["Date", "Account", "Amount", "Memo"])
    ws.append(["2026-03-01", "Revenue", 1000, "Invoice from alice@example.com"])
    ws.append(["2026-03-02", "Refund", -200, "SSN 123-45-6789 on file"])
    ws.append(["2026-03-03", "Commission", -50, "Card 4111-1111-1111-1111"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def merged_cells_xlsx_bytes() -> bytes:
    """xlsx with merged cells on metadata + banner rows.

    Stress-tests the is_merged flag in read_raw_cells — a path the
    real DRONE fixture does not exercise (verified by openpyxl probe:
    every DRONE row had is_merged=False).
    """
    wb = Workbook()
    ws = wb.active
    # Row 1 — merged metadata title
    ws.append(["DRONE Inc. — Profit and Loss — March 2026", None, None, None])
    ws.merge_cells("A1:D1")
    ws["A1"].alignment = Alignment(horizontal="center")

    # Row 2 — header, bold
    ws.append(["Date", "Account", "Amount", "Memo"])
    for cell in ws[2]:
        cell.font = Font(bold=True)

    # Row 3 — merged REVENUE banner, bold
    ws.append([None, "REVENUE", None, None])
    ws.merge_cells("B3:D3")
    for cell in ws[3]:
        if cell.value is not None:
            cell.font = Font(bold=True)

    # Rows 4-5 — line items, no merging
    ws.append(["2026-03-01", "Product Revenue", 1000, "Sales"])
    ws.append(["2026-03-02", "Service Revenue", 500, "Consulting"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
